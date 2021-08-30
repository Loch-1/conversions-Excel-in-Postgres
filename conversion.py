#!/usr/bin/env python3
# -*- encoding: utf-8 -*-

from openpyxl import load_workbook

import psycopg2
import config

ver = "1.0"
copyleft = "(c) 2021 Viktor"
comment = "Скрипт для конвертации таблиц Excel в базу даннух Postgres"

print(f"{ver} \n{copyleft}\n{comment}\n\n")

sourseFileUserUNEP = 'D:\\project\\conversion Excel in Postgres\\!initial data\\UNEP_users.xlsx'
sourseFileUserUKEP = 'D:\\project\\conversion Excel in Postgres\\!initial data\\UKEP_users.xlsx'
sourseFileSertUNEP = 'D:\\project\\conversion Excel in Postgres\\!initial data\\UNEP_certs_Short.xlsx'
sourseFileSertUKEP = 'D:\\project\\conversion Excel in Postgres\\!initial data\\UKEP_certs.xlsx'
file_error = 'D:\\project\\conversion Excel in Postgres\\!initial data\\data_error.txt'

userList = {}
certUnepList = []

db_config = config.db_config

conn = psycopg2.connect(**db_config)
cursor = conn.cursor()

print("Информация о сервере PostgreSQL")
print(conn.get_dsn_parameters(), "\n")

cursor.close()
conn.close()


def createTableUCUser():
    conn = psycopg2.connect(**db_config)
    cursor = conn.cursor()

    try:
        cursor.execute('''CREATE TABLE UCUser
            (ID VARCHAR(50) PRIMARY KEY UNIQUE,
            UPN VARCHAR(50) NOT NULL);''')

        print('Таблица UCUser создана')

    except psycopg2.errors.DuplicateTable:
        print ('Таблица UCUser существует')

    conn.commit()
    cursor.close()
    conn.close()

def createTableCertType():
    conn = psycopg2.connect(**db_config)
    cursor = conn.cursor()

    try:
        cursor.execute('''CREATE TABLE CertificateType
            (ID SERIAL PRIMARY KEY,
            certificateTypeName CHAR(20) UNIQUE,
            CertificateTypeDescription VARCHAR(100));''')

        print('Таблица CertificateType создана')

    except psycopg2.errors.DuplicateTable:
        print ('Таблица CertificateType существует')

    conn.commit()

    _SQL = """INSERT INTO public.CertificateType (certificateTypeName) VALUES ('UNEP')"""
    cursor.execute(_SQL)

    _SQL = """INSERT INTO public.CertificateType (certificateTypeName) VALUES ('UKEP')"""
    cursor.execute(_SQL)
    conn.commit()
    cursor.close()
    conn.close()

    print('CertificateType table is filled in')


def createTableCertificateUNEP():
    conn = psycopg2.connect(**db_config)
    cursor = conn.cursor()

    try:
        cursor.execute('''CREATE TABLE CertificateUNEP
            (ID SERIAL PRIMARY KEY,
            CertificateType int REFERENCES CertificateType(ID) ON DELETE CASCADE,
            Thumbprint VARCHAR(50) NOT NULL,
            validAfter DATE NOT NULL,
            validBefore DATE NOT NULL,            
            IsDefault boolean,
            UCUser VARCHAR(50) REFERENCES UCUser(ID) ON DELETE CASCADE);''')

        print('Create a Certificate table')

    except psycopg2.errors.DuplicateTable:
        print ('Table the Certificate exists')

    conn.commit()
    cursor.close()
    conn.close()

def readUsersUnep():
    wb = load_workbook(sourseFileUserUNEP)
    ws = wb['Лист1']
    i = 0

    for row in range(2, ws.max_row + 1):
        if ws["A" + str(row)].value:
            userID = ws["A" + str(row)].value  # id из файла Эксель
            providerKey = str(ws["B" + str(row)].value)  # Логин из файла Эксель
            i += 1

            if userID not in userList:
                userList[userID] = providerKey
                print(f'--- Записываем значение {i} ----{userID} {providerKey}')

            elif userID in userList and '@' in providerKey:
                print(f'--- Записываем значение {i} ----{userID} {providerKey}')
                userList[userID] = providerKey


'''
Функция проверяет всели данные из таблицы занесены в БД            
def test():
    wb = load_workbook(sourseFileUserUNEP)
    ws = wb['Лист1']
    i = 0
    print("Начало теста")

    for row in range(2, ws.max_row+1):
        if ws["A"+str(row)].value:
            userID = ws["A"+str(row)].value # id из файла Эксель
            providerKey = str(ws["B"+str(row)].value) # Логин из файла Эксель

            if userList[userID].split('@')[0] != providerKey.split('@')[0]:
                print (f'Ошибка {userList[userID]}/{providerKey}')
'''

def readCertificateUNEP():
    wb = load_workbook(sourseFileSertUNEP)
    ws = wb['Лист1']
    i = 0

    for row in range(2, ws.max_row+1):
        i += 1

        if ws["A" + str(row)].value:
            Thumbprint = ws["A" + str(row)].value
            validAfter = ws["B" + str(row)].value
            validBefore = ws["C" + str(row)].value
            IsDefault = ws["D" + str(row)].value
            UCUser = ws["E" + str(row)].value
            print(f'--- Записываем значение {i} ----{Thumbprint}')
            certUnepList.append([Thumbprint, validAfter, validBefore, IsDefault, UCUser])

    for item in certUnepList:
        print (item[0], item[1], item[2], item[3], item[4])


def saveTableUserUNEP():
    i = 0
    conn = psycopg2.connect(**db_config)
    cursor = conn.cursor()
    for UserID, ProviderKey in userList.items():
        i += 1
        print(f'---Пишим в БД {i} значение')
        _SQL = """INSERT INTO public.UCUser(ID, UPN)
            VALUES ('%(UserID)s', '%(ProviderKey)s')
            """ %{'UserID': UserID, 'ProviderKey': ProviderKey}

        cursor.execute(_SQL)
        conn.commit()

    print('Table UCUser write in Postgres')


def saveTableTableCertificateUNEP():
    i = 0
    conn = psycopg2.connect(**db_config)
    cursor = conn.cursor()
    for item in certUnepList:
        i += 1
        _SQL = """INSERT INTO public.CertificateUNEP
        (CertificateType, Thumbprint, validAfter, validBefore, IsDefault, UCUser)
        VALUES
        
        """

        print(item[0], item[1], item[2], item[3], item[4])

        # cursor.execute('''CREATE TABLE CertificateUNEP
        #      (ID SERIAL PRIMARY KEY,
        #      CertificateType int REFERENCES CertificateType(ID) ON DELETE CASCADE,
        #      Thumbprint VARCHAR(50) NOT NULL,
        #      validBefore DATE NOT NULL,
        #      validAfter DATE NOT NULL,
        #      IsDefault boolean,
        #      UCUser VARCHAR(50) REFERENCES UCUser(ID) ON DELETE CASCADE);''')


        # createTableUCUser()
# readUsersUnep()
# saveTableUserUNEP()
# createTableCertType()
# createTableCertificateUNEP()

readCertificateUNEP()
